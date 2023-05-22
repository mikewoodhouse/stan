import warnings

from openpyxl import load_workbook

from app.types.classes import Match


class ExcelLoader:
    def __init__(self, filepath: str) -> None:
        self.filepath = filepath


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

wb = load_workbook("/mnt/c/Users/mikew/Documents/TOCC/Averages/TOCC Averages 2022.xlsm")

sh = wb["Matches"]

hdrs = None
match_rows = []
for row in sh.iter_rows(min_col=1, max_col=30, values_only=True):
    if row[0] == "Day":
        hdrs = row
    elif hdrs and row[1]:
        match_rows.append(row)
    elif hdrs:
        break

match_map = {
    1: "date",
    2: "oppo",
    3: "venue",
    4: "result",
    5: "bat_first",
    6: "first_runs",
    7: "first_wkts",
    8: "first_all_out",
    9: "first_notes",
    10: "second_runs",
    11: "second_wkts",
    12: "second_all_out",
    13: "second_notes",
    14: "overs_opp",
    15: "overs_tocc",
    16: "tocc_w",
    17: "tocc_nb",
    18: "tocc_b",
    19: "tocc_lb",
    20: "opp_w",
    21: "opp_nb",
    22: "opp_b",
    23: "opp_lb",
}

dict_0 = {v: match_rows[3][k] for k, v in match_map.items()}

match = Match(**dict_0)  # type: ignore
